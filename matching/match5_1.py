import os
import re
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext
import threading
import time
import queue
import concurrent.futures
from concurrent.futures import ThreadPoolExecutor
import ctypes
import zipfile
import tempfile
import shutil
import subprocess
import atexit
from copy import copy
import xlwings as xw
import configparser
import pandas as pd
from openpyxl import load_workbook

# 如果是Windows系统且被打包成exe，彻底隐藏控制台窗口
if sys.platform == "win32" and hasattr(sys, 'frozen'):
    ctypes.windll.user32.ShowWindow(ctypes.windll.kernel32.GetConsoleWindow(), 0)
    # 设置子进程不显示控制台窗口
    startupinfo = subprocess.STARTUPINFO()
    startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
    startupinfo.wShowWindow = 0  # SW_HIDE
else:
    startupinfo = None


class PDFExcelTool:
    def __init__(self, root):
        self.root = root
        self.root.title("PDF信息提取与比对工具")

        # 设置DPI感知（仅Windows）
        if sys.platform == "win32":
            try:
                # 启用DPI感知
                ctypes.windll.shcore.SetProcessDpiAwareness(1)
            except (AttributeError, OSError):
                pass  # 如果API不可用则忽略

        # 计算相对尺寸
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        width = int(min(1000, screen_width * 0.8))
        height = int(min(900, screen_height * 0.8))
        self.root.geometry(f"{width}x{height}")
        self.root.minsize(800, 700)  # 最小尺寸
        self.root.resizable(True, True)  # 允许调整大小

        # 居中窗口
        self.center_window()

        # 初始化变量
        self.excel_path = ""
        self.pdf_folder = ""
        self.running = False
        self.report_path = ""
        self.progress_queue = queue.Queue()
        self.log_queue = queue.Queue()
        self.total_pdfs = 0
        self.processed_count = 0
        self.progress_frame = None  # 延迟创建进度条
        self.temp_dir = None  # 存储临时解压目录
        self._7z_path = None  # 存储7z工具路径
        self._7z_dir = None  # 7z工具的临时目录
        self.excel_app = None  # xlwings应用实例
        self.excel_book = None  # xlwings工作簿实例

        # 配置相关变量
        self.config = configparser.ConfigParser()
        self.config_file = "config.ini"
        self.header_row = 23  # 默认表头行
        self.note_start_row = 39  # 默认备注开始行
        self.name_col = 2  # 默认物料名称列
        self.spec_col = 3  # 默认物料规格列
        self.desc_col = 4  # 默认描述列
        self.version_col = 9  # 默认版本列
        self.title_col = 13  # 新增：TITLE对应的列，第13列 (Name and Specification)

        # 注册程序退出时的清理函数
        atexit.register(self.cleanup_on_exit)
        # 当窗口关闭时也执行清理
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

        # 加载配置
        self.load_config()

        # 创建主框架
        self.main_frame = ttk.Frame(self.root, padding=10)
        self.main_frame.pack(fill=tk.BOTH, expand=True)

        # 文件选择部分
        self.create_file_selection_section()

        # 日志输出部分
        self.create_log_section()

        # 状态栏
        self.status_var = tk.StringVar()
        self.status_var.set("就绪")
        self.status_bar = ttk.Label(self.root, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)

        # 启动日志更新线程
        self.log_update_thread = threading.Thread(target=self.update_log, daemon=True)
        self.log_update_thread.start()

        # 打印欢迎信息和配置信息
        welcome_msg = "=" * 70 + "\n"
        welcome_msg += "PDF信息提取与比对工具\n"
        welcome_msg += f"启动时间: {time.strftime('%Y-%m-%d %H:%M:%S')}\n"
        welcome_msg += "支持压缩格式: ZIP, RAR, 7z, TAR, GZ, BZ2等\n"
        welcome_msg += "=" * 70 + "\n"
        welcome_msg += "当前配置:\n"
        welcome_msg += f"  表头行: {self.header_row}\n"
        welcome_msg += f"  备注开始行: {self.note_start_row}\n"
        welcome_msg += f"  物料名称列: {self.name_col}\n"
        welcome_msg += f"  物料规格列: {self.spec_col}\n"
        welcome_msg += f"  描述列: {self.desc_col}\n"
        welcome_msg += f"  版本列: {self.version_col}\n"
        welcome_msg += f"  TITLE列 (Name and Specification): {self.title_col}\n"  # 新增
        welcome_msg += "=" * 70 + "\n"
        welcome_msg += "使用说明:\n"
        welcome_msg += "1. 选择PDF图纸文件夹或压缩包\n"
        welcome_msg += "2. 点击'开始填入'按钮选择Excel文件并开始提取和填充\n"
        welcome_msg += "3. 点击'开始比对'按钮选择Excel文件并开始比对\n"
        welcome_msg += "4. 处理完成后可点击'查看报告'查看详细结果\n"
        welcome_msg += "=" * 70 + "\n"
        self.log_queue.put(welcome_msg)

        # 初始化7z工具（如果是打包版本）
        if hasattr(sys, 'frozen'):
            self.init_7z_tool()

    def load_config(self):
        """加载配置文件"""
        # 默认配置
        default_config = {
            'EXCEL': {
                'header_row': '23',
                'note_start_row': '39',
                'name_col': '2',
                'spec_col': '3',
                'desc_col': '4',
                'version_col': '9',
                'title_col': '13'
            }
        }

        # 如果配置文件不存在，创建默认配置
        if not os.path.exists(self.config_file):
            self.config.read_dict(default_config)
            with open(self.config_file, 'w', encoding='utf-8') as f:
                self.config.write(f)
            self.log_queue.put("创建默认配置文件\n")
        else:
            # 读取现有配置
            self.config.read(self.config_file, encoding='utf-8')

        # 更新配置变量
        try:
            self.header_row = int(self.config['EXCEL']['header_row'])
            self.note_start_row = int(self.config['EXCEL']['note_start_row'])
            self.name_col = int(self.config['EXCEL']['name_col'])
            self.spec_col = int(self.config['EXCEL']['spec_col'])
            self.desc_col = int(self.config['EXCEL']['desc_col'])
            self.version_col = int(self.config['EXCEL']['version_col'])
            self.title_col = int(self.config['EXCEL']['title_col'])  # 新增：读取TITLE列配置
            self.log_queue.put("配置文件加载成功\n")
        except (KeyError, ValueError) as e:
            self.log_queue.put(f"配置加载错误: {str(e)}，使用默认值\n")

    def init_7z_tool(self):
        """初始化7z工具（支持重复调用，确保工具文件存在）"""
        # 如果已存在有效路径，直接返回
        if self._7z_path and os.path.exists(self._7z_path):
            return True

        try:
            # 若之前的临时目录无效，创建新的临时目录
            if self._7z_dir and (not os.path.exists(self._7z_dir)):
                self._7z_dir = None  # 标记为无效

            if not self._7z_dir:
                self._7z_dir = tempfile.mkdtemp(prefix="7z_")
                self.log_queue.put(f"创建7z工具临时目录: {self._7z_dir}\n")

            self._7z_path = os.path.join(self._7z_dir, "7z.exe")

            # 从打包资源中提取7z.exe和7z.dll
            if hasattr(sys, '_MEIPASS'):
                # 检查资源中是否存在7z文件
                resource_exe = os.path.join(sys._MEIPASS, "7z.exe")
                resource_dll = os.path.join(sys._MEIPASS, "7z.dll")

                if not os.path.exists(resource_exe):
                    self.log_queue.put("错误：未在打包资源中找到7z.exe\n")
                    return False

                # 复制文件到临时目录（如果不存在或已损坏）
                if not os.path.exists(self._7z_path) or os.path.getsize(self._7z_path) == 0:
                    shutil.copy(resource_exe, self._7z_path)
                    self.log_queue.put(f"已提取7z.exe到临时目录\n")

                # 处理7z.dll
                dll_dest = os.path.join(self._7z_dir, "7z.dll")
                if os.path.exists(resource_dll) and (not os.path.exists(dll_dest) or os.path.getsize(dll_dest) == 0):
                    shutil.copy(resource_dll, dll_dest)
                    self.log_queue.put(f"已提取7z.dll到临时目录\n")

                self.log_queue.put(f"7z工具已初始化: {self._7z_path}\n")
                return True
            else:
                self.log_queue.put("警告：未在打包环境中运行，使用系统7z工具\n")
                # 尝试使用系统7z
                self._7z_path = "7z"
                return True
        except Exception as e:
            self.log_queue.put(f"初始化7z工具失败: {str(e)}\n")
            return False

    def center_window(self):
        """居中窗口"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f"+{x}+{y}")

    def create_file_selection_section(self):
        """创建文件选择区域"""
        file_frame = ttk.LabelFrame(self.main_frame, text="文件选择", padding=10)
        file_frame.pack(fill=tk.X, pady=(0, 10))

        # PDF文件夹/压缩包选择
        pdf_frame = ttk.Frame(file_frame)
        pdf_frame.grid(row=0, column=0, sticky=tk.EW, padx=5, pady=5)
        ttk.Label(pdf_frame, text="PDF图纸文件夹/压缩包:").pack(side=tk.LEFT)
        self.pdf_entry = ttk.Entry(pdf_frame, width=80)
        self.pdf_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 0))
        ttk.Button(pdf_frame, text="浏览...", command=self.select_pdf_folder_or_archive, width=10).pack(side=tk.LEFT,
                                                                                                        padx=(5, 0))

        # 按钮区域
        btn_frame = ttk.Frame(file_frame)
        btn_frame.grid(row=1, column=0, columnspan=3, pady=10, sticky=tk.EW)

        # 使用Frame作为容器，使按钮居中
        center_frame = ttk.Frame(btn_frame)
        center_frame.pack(expand=True)

        self.fill_btn = ttk.Button(center_frame, text="开始填入", command=self.start_filling, width=15)
        self.fill_btn.pack(side=tk.LEFT, padx=5)

        self.compare_btn = ttk.Button(center_frame, text="开始比对", command=self.start_comparison, width=15)
        self.compare_btn.pack(side=tk.LEFT, padx=5)

        self.report_btn = ttk.Button(center_frame, text="查看报告", command=self.open_report, state=tk.DISABLED,
                                     width=15)
        self.report_btn.pack(side=tk.LEFT, padx=5)

        ttk.Button(center_frame, text="退出", command=self.on_close, width=15).pack(side=tk.LEFT, padx=5)

        # 配置网格权重
        file_frame.columnconfigure(0, weight=1)

    def create_log_section(self):
        """创建日志输出区域"""
        log_frame = ttk.LabelFrame(self.main_frame, text="处理日志", padding=10)
        log_frame.pack(fill=tk.BOTH, expand=True)

        self.log_text = scrolledtext.ScrolledText(
            log_frame,
            wrap=tk.WORD,
            font=("Consolas", 10),
            bg="#f0f0f0",
            padx=10,
            pady=10
        )
        self.log_text.pack(fill=tk.BOTH, expand=True)
        self.log_text.config(state=tk.DISABLED)  # 初始禁用编辑

    def update_log(self):
        """后台线程更新日志显示"""
        while True:
            try:
                # 从队列获取日志消息
                msg = self.log_queue.get(timeout=0.1)

                # 更新日志文本框
                self.log_text.config(state=tk.NORMAL)
                self.log_text.insert(tk.END, msg)
                self.log_text.see(tk.END)
                self.log_text.config(state=tk.DISABLED)

                # 更新GUI
                self.root.update_idletasks()
            except queue.Empty:
                time.sleep(0.05)
            except Exception:
                pass

    def create_progress_section(self):
        """创建进度条区域（在需要时创建）"""
        if self.progress_frame is None:
            self.progress_frame = ttk.Frame(self.main_frame)
            self.progress_frame.pack(fill=tk.X, pady=(0, 10))

            # 进度条标签
            self.progress_label = ttk.Label(self.progress_frame, text="进度: 0%")
            self.progress_label.pack(side=tk.LEFT, padx=5)

            # 进度条
            self.progress_bar = ttk.Progressbar(
                self.progress_frame,
                orient=tk.HORIZONTAL,
                length=600,
                mode='determinate'
            )
            self.progress_bar.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)

            # 百分比标签
            self.percent_label = ttk.Label(self.progress_frame, text="0%")
            self.percent_label.pack(side=tk.LEFT, padx=5)
        else:
            # 重置进度条
            self.progress_bar['value'] = 0
            self.progress_label.config(text="进度: 0%")
            self.percent_label.config(text="0%")

    def select_pdf_folder_or_archive(self):
        """选择PDF文件、压缩包或文件夹"""
        # 创建选择对话框
        choice = tk.messagebox.askquestion(
            "选择类型",
            "请选择要添加的内容类型：\n\n'是' - PDF文件或压缩包\n'否' - PDF文件夹",
            icon='question'
        )

        if choice == 'yes':
            # 选择PDF文件或压缩包
            file_path = filedialog.askopenfilename(
                title="选择PDF文件或压缩包",
                filetypes=[
                    ("压缩文件", "*.zip *.rar *.7z *.tar *.gz *.bz2"),
                    ("PDF文件", "*.pdf"),
                    ("所有文件", "*.*")
                ]
            )
            if file_path:
                self.pdf_folder = file_path
                self.pdf_entry.delete(0, tk.END)
                self.pdf_entry.insert(0, file_path)
        else:
            # 选择PDF文件夹
            folder_path = filedialog.askdirectory(title="选择PDF图纸文件夹")
            if folder_path:
                self.pdf_folder = folder_path
                self.pdf_entry.delete(0, tk.END)
                self.pdf_entry.insert(0, folder_path)

    def start_filling(self):
        """开始填充过程"""
        if not self.pdf_folder:
            messagebox.showerror("错误", "请先选择PDF文件夹或压缩包")
            return

        if not os.path.exists(self.pdf_folder):
            messagebox.showerror("错误", f"路径不存在: {self.pdf_folder}")
            return

        # 选择Excel文件
        excel_path = filedialog.askopenfilename(
            title="选择要填充的Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls *.xlsm"), ("所有文件", "*.*")]
        )

        if not excel_path:
            return

        self.excel_path = excel_path

        # 重新加载配置
        self.load_config()
        self.log_queue.put("开始处理前已重新加载配置\n")
        # 显示当前配置
        config_msg = "当前使用的配置:\n"
        config_msg += f"  表头行: {self.header_row}\n"
        config_msg += f"  备注开始行: {self.note_start_row}\n"
        config_msg += f"  物料名称列: {self.name_col}\n"
        config_msg += f"  物料规格列: {self.spec_col}\n"
        config_msg += f"  描述列: {self.desc_col}\n"
        config_msg += f"  版本列: {self.version_col}\n"
        config_msg += f"  TITLE列 (Name and Specification): {self.title_col}\n"  # 新增
        self.log_queue.put(config_msg)

        # 创建进度条区域
        self.create_progress_section()

        # 禁用按钮
        self.fill_btn.config(state=tk.DISABLED)
        self.compare_btn.config(state=tk.DISABLED)
        self.report_btn.config(state=tk.DISABLED)
        self.running = True
        self.status_var.set("正在处理中...")

        # 清空日志
        self.log_text.config(state=tk.NORMAL)
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state=tk.DISABLED)

        # 在新线程中运行处理过程
        threading.Thread(target=self.run_filling, daemon=True).start()

        # 启动进度更新
        self.root.after(100, self.update_progress)

    def start_comparison(self):
        """开始比对过程"""
        if not self.pdf_folder:
            messagebox.showerror("错误", "请先选择PDF文件夹或压缩包")
            return

        if not os.path.exists(self.pdf_folder):
            messagebox.showerror("错误", f"路径不存在: {self.pdf_folder}")
            return

        # 选择Excel文件
        excel_path = filedialog.askopenfilename(
            title="选择要比对的Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )

        if not excel_path:
            return

        self.excel_path = excel_path

        # 创建进度条区域
        self.create_progress_section()

        # 禁用按钮
        self.fill_btn.config(state=tk.DISABLED)
        self.compare_btn.config(state=tk.DISABLED)
        self.report_btn.config(state=tk.DISABLED)
        self.running = True
        self.status_var.set("正在比对中...")

        # 清空日志
        self.log_text.config(state=tk.NORMAL)
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state=tk.DISABLED)

        # 在新线程中运行比对过程
        threading.Thread(target=self.run_comparison, daemon=True).start()

        # 启动进度更新
        self.root.after(100, self.update_progress)

    def update_progress(self):
        """更新进度条显示"""
        if not self.running:
            return

        try:
            # 从队列中获取进度更新
            while True:
                processed_count = self.progress_queue.get_nowait()
                self.processed_count = processed_count

                # 计算进度百分比
                if self.total_pdfs > 0:
                    progress = (processed_count / self.total_pdfs) * 100
                    self.progress_bar['value'] = progress
                    percent_text = f"{int(progress)}%"
                    self.progress_label.config(text=f"进度: {percent_text}")
                    self.percent_label.config(text=percent_text)
        except queue.Empty:
            pass

        # 继续调度下一次更新
        if self.running:
            self.root.after(100, self.update_progress)

    def extract_archive(self, archive_path, extract_dir):
        """使用7z工具解压压缩包"""
        # 每次解压前检查7z工具是否有效，无效则重新初始化
        if not self.init_7z_tool():
            self.log_queue.put("7z工具不可用，无法解压文件\n")
            return False

        try:
            # 构建解压命令
            if self._7z_path and os.path.exists(self._7z_path):
                cmd = [self._7z_path, "x", archive_path, f"-o{extract_dir}", "-y"]
            else:
                # 回退到系统7z（如果有）
                cmd = ["7z", "x", archive_path, f"-o{extract_dir}", "-y"]

            # 运行解压命令
            result = subprocess.run(
                cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == "win32" else 0,
                startupinfo=startupinfo
            )

            if result.returncode != 0:
                error_msg = f"解压失败 (错误码 {result.returncode}): {result.stderr.decode('gbk', errors='ignore')}"
                self.log_queue.put(error_msg + "\n")
                return False

            return True
        except Exception as e:
            self.log_queue.put(f"解压过程中出错: {str(e)}\n")
            return False

    def run_filling(self):
        """执行填充过程"""
        try:
            # 添加时间戳
            start_msg = "=" * 70 + "\n"
            start_msg += f"PDF信息提取并填充到Excel工具 - 开始时间: {time.strftime('%Y-%m-%d %H:%M:%S')}\n"
            start_msg += "=" * 70 + "\n"
            self.log_queue.put(start_msg)

            # 处理压缩包（如果是压缩文件）
            actual_pdf_folder = self.pdf_folder
            is_archive = False
            archive_type = ""

            # 支持的压缩文件扩展名
            archive_extensions = ['.zip', '.rar', '.7z', '.tar', '.gz', '.bz2']

            # 检查是否是压缩文件
            if os.path.isfile(self.pdf_folder):
                file_ext = os.path.splitext(self.pdf_folder)[1].lower()

                # 如果是PDF文件，直接使用
                if file_ext == '.pdf':
                    actual_pdf_folder = os.path.dirname(self.pdf_folder)
                # 如果是支持的压缩格式
                elif file_ext in archive_extensions:
                    archive_type = file_ext[1:].upper()  # 去掉点，转换为大写
                    self.log_queue.put(f"检测到{archive_type}压缩包: {self.pdf_folder}\n")
                    self.log_queue.put("正在解压缩...\n")

                    # 创建临时目录
                    self.temp_dir = tempfile.mkdtemp()
                    self.log_queue.put(f"创建临时目录: {self.temp_dir}\n")

                    try:
                        # 使用7z工具解压
                        success = self.extract_archive(self.pdf_folder, self.temp_dir)
                        if not success:
                            error_msg = f"解压{archive_type}文件失败\n"
                            self.log_queue.put(error_msg)
                            self.root.after(0, lambda: messagebox.showerror(
                                "解压错误",
                                f"无法解压{archive_type}文件\n请确保压缩包未损坏"
                            ))
                            self.complete_processing()
                            return

                        actual_pdf_folder = self.temp_dir
                        is_archive = True
                        self.log_queue.put(f"{archive_type}解压缩完成!\n")
                    except Exception as e:
                        error_msg = f"解压{archive_type}文件失败: {str(e)}\n"
                        self.log_queue.put(error_msg)
                        self.root.after(0, lambda: messagebox.showerror(
                            "解压错误",
                            f"无法解压{archive_type}文件:\n{str(e)}"
                        ))
                        self.complete_processing()
                        return
                else:
                    error_msg = f"不支持的文件格式: {file_ext}\n"
                    self.log_queue.put(error_msg)
                    self.root.after(0, lambda: messagebox.showerror(
                        "错误",
                        f"不支持的文件格式: {file_ext}"
                    ))
                    self.complete_processing()
                    return

            # 获取PDF文件列表（递归搜索）
            self.log_queue.put(f"正在搜索PDF文件: {actual_pdf_folder}\n")
            pdf_files = self.find_pdf_files(actual_pdf_folder)
            self.total_pdfs = len(pdf_files)

            if self.total_pdfs == 0:
                warn_msg = f"警告: 路径中没有找到PDF文件 - {actual_pdf_folder}\n"
                self.log_queue.put(warn_msg)
                self.root.after(0, lambda: messagebox.showwarning(
                    "警告",
                    f"路径中没有找到PDF文件:\n{actual_pdf_folder}"
                ))
                self.complete_processing()
                return

            self.log_queue.put(f"找到 {self.total_pdfs} 个PDF文件\n")

            # 打开Excel文件（使用xlwings）
            try:
                self.log_queue.put(f"打开Excel文件: {self.excel_path}\n")
                self.excel_app = xw.App(visible=False)  # 隐藏Excel窗口
                self.excel_book = self.excel_app.books.open(self.excel_path)

                # 检查表头行和备注行之间的可用行数
                sheet = self.excel_book.sheets.active
                available_rows = self.note_start_row - self.header_row - 1  # 表头和备注行之间的行数

                self.log_queue.put(
                    f"检测到表头行({self.header_row})和备注行({self.note_start_row})之间可用行数: {available_rows}\n")
                self.log_queue.put(f"需要处理的PDF文件数: {self.total_pdfs}\n")

                # 如果可用行数不足，插入新行
                if available_rows < self.total_pdfs:
                    need_rows = self.total_pdfs - available_rows
                    self.log_queue.put(f"可用行数不足，需要插入 {need_rows} 行\n")

                    # 在表头行之后插入新行
                    insert_position = self.header_row + 2  # 在表头行后插入
                    sheet.api.Rows(f"{insert_position}:{insert_position + need_rows - 1}").Insert()
                    self.log_queue.put(f"已在第{insert_position}行位置插入 {need_rows} 行\n")

                    # 更新备注行的位置
                    self.note_start_row += need_rows
                    self.log_queue.put(f"备注行已更新到第 {self.note_start_row} 行\n")

            except Exception as e:
                error_msg = f"打开Excel文件失败: {str(e)}\n"
                self.log_queue.put(error_msg)
                self.root.after(0, lambda: messagebox.showerror(
                    "Excel错误",
                    f"无法打开Excel文件:\n{str(e)}"
                ))
                self.complete_processing()
                return

            # 处理文件
            self.log_queue.put(f"开始处理 {self.total_pdfs} 个PDF文件...\n")
            results = self.process_files_for_filling(
                self.excel_path,
                actual_pdf_folder,
                self.progress_queue,
                self.log_queue,
                pdf_files,
                self.excel_book,
                self.header_row,
                self.note_start_row,
                self.name_col,
                self.spec_col,
                self.desc_col,
                self.version_col,
                self.title_col  # 新增：传递TITLE列参数
            )

            # 保存Excel文件到excel文件夹
            try:
                # 创建excel文件夹（如果不存在）
                if not os.path.exists('excel'):
                    os.makedirs('excel')
                    self.log_queue.put("已创建excel文件夹\n")

                # 获取原文件名并添加时间戳
                original_filename = os.path.basename(self.excel_path)
                name, ext = os.path.splitext(original_filename)
                timestamp = time.strftime('%Y%m%d_%H%M%S')
                new_filename = f"{name}_{timestamp}{ext}"
                new_filepath = os.path.join('excel', new_filename)

                self.excel_book.save(new_filepath)
                self.log_queue.put(f"已另存Excel文件到: {new_filepath}\n")
            except Exception as e:
                self.log_queue.put(f"保存Excel文件时出错: {str(e)}\n")

            # 创建log文件夹（如果不存在的话）
            if not os.path.exists('log'):
                os.makedirs('log')

            # 保存文件的路径
            report_file = os.path.join('log', f"处理报告_{time.strftime('%Y%m%d_%H%M%S')}.txt")
            self.report_path = self.generate_filling_report(results, report_file)

            # 显示完成消息
            complete_msg = "\n" + "=" * 70 + "\n"
            complete_msg += f"处理完成！已处理 {len(results)} 个PDF文件。\n"

            # 统计多页PDF数量
            multi_page_count = sum(1 for result in results if result['page_count'] > 1)
            if multi_page_count > 0:
                complete_msg += f"发现 {multi_page_count} 个多页PDF文件，需要进一步查看。\n"

            complete_msg += f"报告已保存到: {self.report_path}\n"
            self.log_queue.put(complete_msg)

            # 启用报告按钮
            self.report_btn.config(state=tk.NORMAL)
            self.running = False

            # 弹出完成消息
            self.root.after(0, lambda: messagebox.showinfo(
                "处理完成",
                f"处理完成！已处理 {len(results)} 个PDF文件。\n\n" +
                (f"发现 {multi_page_count} 个多页PDF文件，需要进一步查看。\n" if multi_page_count > 0 else "") +
                f"报告已保存到:\n{self.report_path}"
            ))

        except Exception as e:
            error_msg = f"处理过程中发生错误: {str(e)}\n"
            self.log_queue.put(error_msg)
            self.root.after(0, lambda: messagebox.showerror(
                "错误",
                f"处理过程中发生错误:\n{str(e)}"
            ))
        finally:
            # 关闭Excel
            if self.excel_book:
                try:
                    self.excel_book.close()
                except Exception as e:
                    self.log_queue.put(f"关闭Excel文件时出错: {str(e)}\n")
            if self.excel_app:
                try:
                    self.excel_app.quit()
                except Exception as e:
                    self.log_queue.put(f"退出Excel应用时出错: {str(e)}\n")

            # 清理临时目录（如果是解压的）
            if self.temp_dir and os.path.exists(self.temp_dir):
                try:
                    shutil.rmtree(self.temp_dir)
                    self.log_queue.put(f"已清理临时目录: {self.temp_dir}\n")
                except Exception as e:
                    self.log_queue.put(f"清理临时目录失败: {str(e)}\n")
                self.temp_dir = None

            # 更新状态
            self.complete_processing()

    def run_comparison(self):
        """执行比对过程"""
        try:
            # 添加时间戳
            start_msg = "=" * 70 + "\n"
            start_msg += f"Excel-PDF图纸信息比对工具 - 开始时间: {time.strftime('%Y-%m-%d %H:%M:%S')}\n"
            start_msg += "=" * 70 + "\n"
            self.log_queue.put(start_msg)

            # 处理压缩包（如果是压缩文件）
            actual_pdf_folder = self.pdf_folder
            is_archive = False
            archive_type = ""

            # 支持的压缩文件扩展名
            archive_extensions = ['.zip', '.rar', '.7z', '.tar', '.gz', '.bz2']

            # 检查是否是压缩文件
            if os.path.isfile(self.pdf_folder):
                file_ext = os.path.splitext(self.pdf_folder)[1].lower()

                # 如果是PDF文件，直接使用
                if file_ext == '.pdf':
                    actual_pdf_folder = os.path.dirname(self.pdf_folder)
                # 如果是支持的压缩格式
                elif file_ext in archive_extensions:
                    archive_type = file_ext[1:].upper()  # 去掉点，转换为大写
                    self.log_queue.put(f"检测到{archive_type}压缩包: {self.pdf_folder}\n")
                    self.log_queue.put("正在解压缩...\n")

                    # 创建临时目录
                    self.temp_dir = tempfile.mkdtemp()
                    self.log_queue.put(f"创建临时目录: {self.temp_dir}\n")

                    try:
                        # 使用7z工具解压
                        success = self.extract_archive(self.pdf_folder, self.temp_dir)
                        if not success:
                            error_msg = f"解压{archive_type}文件失败\n"
                            self.log_queue.put(error_msg)
                            self.root.after(0, lambda: messagebox.showerror(
                                "解压错误",
                                f"无法解压{archive_type}文件\n请确保压缩包未损坏"
                            ))
                            self.complete_processing()
                            return

                        actual_pdf_folder = self.temp_dir
                        is_archive = True
                        self.log_queue.put(f"{archive_type}解压缩完成!\n")
                    except Exception as e:
                        error_msg = f"解压{archive_type}文件失败: {str(e)}\n"
                        self.log_queue.put(error_msg)
                        self.root.after(0, lambda: messagebox.showerror(
                            "解压错误",
                            f"无法解压{archive_type}文件:\n{str(e)}"
                        ))
                        self.complete_processing()
                        return
                else:
                    error_msg = f"不支持的文件格式: {file_ext}\n"
                    self.log_queue.put(error_msg)
                    self.root.after(0, lambda: messagebox.showerror(
                        "错误",
                        f"不支持的文件格式: {file_ext}"
                    ))
                    self.complete_processing()
                    return

            # 获取PDF文件列表（递归搜索）
            self.log_queue.put(f"正在搜索PDF文件: {actual_pdf_folder}\n")
            pdf_files = self.find_pdf_files(actual_pdf_folder)
            self.total_pdfs = len(pdf_files)

            if self.total_pdfs == 0:
                warn_msg = f"警告: 路径中没有找到PDF文件 - {actual_pdf_folder}\n"
                self.log_queue.put(warn_msg)
                self.root.after(0, lambda: messagebox.showwarning(
                    "警告",
                    f"路径中没有找到PDF文件:\n{actual_pdf_folder}"
                ))
                self.complete_processing()
                return

            self.log_queue.put(f"找到 {self.total_pdfs} 个PDF文件\n")

            # 提取Excel数据
            self.log_queue.put("读取Excel数据...\n")
            excel_data = self.extract_excel_data(self.excel_path)
            if excel_data.empty:
                self.log_queue.put("错误: 未找到有效的Excel数据\n")
                self.root.after(0, lambda: messagebox.showerror(
                    "错误",
                    "未找到有效的Excel数据"
                ))
                self.complete_processing()
                return

            # 构建Excel索引
            self.log_queue.put("构建Excel数据索引...\n")
            excel_index = self.build_excel_index(excel_data)

            # 处理文件
            self.log_queue.put(f"开始处理 {self.total_pdfs} 个PDF文件...\n")
            errors = self.process_files_for_comparison(
                self.excel_path,
                actual_pdf_folder,
                self.progress_queue,
                self.log_queue,
                pdf_files,
                excel_data,
                excel_index
            )

            # 创建log文件夹（如果不存在的话）
            if not os.path.exists('log'):
                os.makedirs('log')

            # 保存文件的路径
            report_file = os.path.join('log', f"比对报告_{time.strftime('%Y%m%d_%H%M%S')}.txt")
            self.report_path = self.generate_comparison_report(errors, report_file)

            # 显示完成消息
            complete_msg = "\n" + "=" * 70 + "\n"
            if errors:
                result_msg = f"比对完成！发现 {len(errors)} 个错误。\n"
                complete_msg += result_msg + "错误摘要:\n"
                for i, error in enumerate(errors[:5], 1):  # 最多显示前5个错误
                    # 显示Excel行号
                    excel_row_info = f"Excel行: {error['excel_row']}" if error['excel_row'] != "无" else "未匹配到Excel行"
                    complete_msg += f"  {i}. {error['pdf_file']} - {excel_row_info} - {error['match_type']}\n"
                    for err in error['errors']:
                        complete_msg += f"      - {err}\n"
                if len(errors) > 5:
                    complete_msg += f"  还有 {len(errors) - 5} 个错误未显示...\n"
            else:
                result_msg = "恭喜！所有数据比对一致！\n"
                complete_msg += result_msg

            complete_msg += f"报告已保存到: {self.report_path}\n"
            self.log_queue.put(complete_msg)

            # 启用报告按钮
            self.report_btn.config(state=tk.NORMAL)

            # 弹出完成消息
            self.root.after(0, lambda: messagebox.showinfo(
                "比对完成",
                f"{result_msg}\n\n报告已保存到:\n{self.report_path}"
            ))

        except Exception as e:
            error_msg = f"处理过程中发生错误: {str(e)}\n"
            self.log_queue.put(error_msg)
            self.root.after(0, lambda: messagebox.showerror(
                "错误",
                f"处理过程中发生错误:\n{str(e)}"
            ))
        finally:
            # 清理临时目录（如果是解压的）
            if self.temp_dir and os.path.exists(self.temp_dir):
                try:
                    shutil.rmtree(self.temp_dir)
                    self.log_queue.put(f"已清理临时目录: {self.temp_dir}\n")
                except Exception as e:
                    self.log_queue.put(f"清理临时目录失败: {str(e)}\n")
                self.temp_dir = None

            # 更新状态
            self.complete_processing()

    def find_pdf_files(self, folder_path):
        """递归查找文件夹中的所有PDF文件"""
        pdf_files = []

        # 如果是单个PDF文件
        if os.path.isfile(folder_path) and folder_path.lower().endswith('.pdf'):
            return [folder_path]

        # 递归搜索文件夹
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                if file.lower().endswith('.pdf'):
                    pdf_files.append(os.path.join(root, file))
        return pdf_files

    def complete_processing(self):
        """完成处理后的清理工作"""
        self.status_var.set("就绪")
        self.fill_btn.config(state=tk.NORMAL)
        self.compare_btn.config(state=tk.NORMAL)
        self.running = False

        # 确保进度条显示100%
        if self.progress_bar:
            self.progress_bar['value'] = 100
            self.progress_label.config(text="进度: 100%")
            self.percent_label.config(text="100%")

    def open_report(self):
        """打开报告文件"""
        if self.report_path and os.path.isfile(self.report_path):
            try:
                if sys.platform == "win32":
                    os.startfile(self.report_path)
                elif sys.platform == "darwin":  # macOS
                    os.system(f'open "{self.report_path}"')
                else:  # linux
                    os.system(f'xdg-open "{self.report_path}"')
            except Exception as e:
                messagebox.showerror("错误", f"无法打开报告文件:\n{str(e)}")
        else:
            messagebox.showwarning("警告", "报告文件不存在或尚未生成")

    def cleanup_on_exit(self):
        """程序退出时清理临时文件"""
        # 关闭Excel
        if hasattr(self, 'excel_book') and self.excel_book:
            try:
                self.excel_book.close()
            except:
                pass
        if hasattr(self, 'excel_app') and self.excel_app:
            try:
                self.excel_app.quit()
            except:
                pass

        # 清理7z工具的临时目录
        if hasattr(self, '_7z_dir') and self._7z_dir and os.path.exists(self._7z_dir):
            try:
                shutil.rmtree(self._7z_dir)
                print(f"已清理7z工具临时目录: {self._7z_dir}")
            except Exception as e:
                print(f"清理7z工具临时目录失败: {str(e)}")

        # 清理可能残留的PDF解压临时目录
        if hasattr(self, 'temp_dir') and self.temp_dir and os.path.exists(self.temp_dir):
            try:
                shutil.rmtree(self.temp_dir)
                print(f"已清理PDF解压临时目录: {self.temp_dir}")
            except Exception as e:
                print(f"清理PDF解压临时目录失败: {str(e)}")

    def on_close(self):
        """窗口关闭时的处理"""
        # 如果正在运行，先停止
        if self.running:
            self.running = False
            time.sleep(0.5)  # 给线程一点时间停止

        # 执行清理
        self.cleanup_on_exit()

        # 关闭窗口
        self.root.destroy()

    # ======================== 填充功能函数 ========================

    def extract_pdf_title_block(self, pdf_path):
        """从PDF文件中提取标题块信息，优化加工字段提取逻辑（保留中英文）"""
        import pdfplumber
        import re

        title_data = {
            "名称": "", "图号": "", "加工": "", "材料": "", "颜色": "", "表面处理": "", "版本": "", "title": "",
            "页数": 0
        }

        try:
            with pdfplumber.open(pdf_path) as pdf:
                title_data["页数"] = len(pdf.pages)
                if len(pdf.pages) == 0:
                    return title_data

                first_page = pdf.pages[0]
                width = first_page.width
                height = first_page.height
                bbox = (width * 0.3, height * 0.6, width, height)
                cropped_page = first_page.crop(bbox)
                tables = cropped_page.extract_tables()
                extracted_text = cropped_page.extract_text() or ""

                # 单元格合并函数（保持不变）
                def merge_split_cells(table):
                    merged_table = []
                    for row in table:
                        merged_row = []
                        i = 0
                        while i < len(row):
                            cell = str(row[i]).strip()
                            if i + 1 < len(row) and re.match(r'^[a-zA-Z]+\d+$', cell) and re.match(r'^t=[\d.]+$', str(
                                    row[i + 1]).strip()):
                                merged_cell = f"{cell} {str(row[i + 1]).strip()}"
                                merged_row.append(merged_cell)
                                i += 2
                            else:
                                merged_row.append(cell)
                                i += 1
                        merged_table.append(merged_row)
                    return merged_table

                # 表格搜索函数（保持不变）
                def find_in_grid(keywords, search_range=5, ignore_values=[]):
                    matches = []
                    for r, row in enumerate(table_grid):
                        for c, cell in enumerate(row):
                            clean_cell = re.sub(r'\s+', '', cell).lower()
                            for keyword in keywords:
                                if re.sub(r'\s+', '', keyword).lower() in clean_cell:
                                    for i in range(1, search_range + 1):
                                        if c + i < len(row):
                                            value = row[c + i].strip()
                                            if value and value.lower() not in [v.lower() for v in ignore_values]:
                                                matches.append((value, r))
                    if matches:
                        matches.sort(key=lambda x: x[1], reverse=True)
                        return matches[0][0]
                    return ""

                # 处理表格（保持不变）
                table_grid = []
                if tables:
                    merged_tables = [merge_split_cells(table) for table in tables]
                    for table in merged_tables:
                        for row in table:
                            clean_row = [str(cell).strip() if cell is not None else "" for cell in row]
                            table_grid.append(clean_row)

                # 其他字段提取（保持不变）
                ignore_list = ["none", "无", "空白", "/", ""]
                title_data["名称"] = find_in_grid(["名称", "name"], ignore_values=ignore_list)
                title_data["图号"] = find_in_grid(["图号", "图名", "drawing", "DWG NO."], ignore_values=ignore_list)
                title_data["材料"] = find_in_grid(["材料", "material"], ignore_values=ignore_list)
                title_data["颜色"] = find_in_grid(["颜色", "color"], ignore_values=ignore_list)
                title_data["表面处理"] = find_in_grid(["表面处理", "表面", "surface", "SURFACE\nFINISHING"],
                                                      ignore_values=ignore_list)

                # -------------------------- 优化加工字段提取（保留中英文） --------------------------
                # 1. 提取原始加工信息
                processing_raw = find_in_grid(
                    ["加工", "processing", "processes", "MANUFACTUIING\nPROCESSES"],
                    ignore_values=ignore_list
                )

                # 2. 清洗逻辑：保留所有有效内容（中英文），只去除末尾无意义后缀
                if processing_raw:
                    # 定义需要去除的末尾无意义词汇（可扩展）
                    suffix_to_remove = r'(中|中文|了|的|等|完毕|完成|结束)$'

                    # 先去除首尾空格
                    processed = processing_raw.strip()

                    # 循环去除末尾的无意义后缀（可能有多个）
                    while re.search(suffix_to_remove, processed):
                        processed = re.sub(suffix_to_remove, '', processed).strip()

                    title_data["加工"] = processed
                else:
                    title_data["加工"] = ""
                # ------------------------------------------------------------------

                # 版本提取（保持之前的优化逻辑）
                # 关键修改：find_in_grid增加search_range=6
                version_raw = find_in_grid(
                    ["版本", "版 本", "version", "rev"],
                    ignore_values=ignore_list,
                    search_range=6  # 扩大搜索范围，覆盖更多拆分场景
                )

                # 2. 正则匹配：支持小数点前后空格，容错性更强
                # 优化后正则：允许V/Rev后有空格、小数点前后有空格（如V0 .1 → 匹配后清理为V0.1）
                version_pattern = r'(V|Rev|rev)\.?\s*(\d+(?:\s*\.\s*\d+)*)([a-zA-Z]*)'
                version_match = re.search(version_pattern, version_raw, re.IGNORECASE)

                table_extracted = ""  # 存储表格提取的版本号
                if version_match:
                    prefix = version_match.group(1).strip()
                    number_part = version_match.group(2).strip()
                    suffix = version_match.group(3).strip()

                    # 清理数字部分的空格（如“0 .1”→“0.1”，“0. 1”→“0.1”）
                    number_part = re.sub(r'\s*\.\s*', '.', number_part)
                    # 清理SIZE关键词（保留原有逻辑）
                    if 'SIZE' in suffix.upper():
                        size_pos = suffix.upper().find('SIZE')
                        suffix = suffix[:size_pos]
                    table_extracted = f"{prefix}{number_part}{suffix}"

                # 3. 二次校验：若表格提取结果是“短版本号”（如V0、Rev1），强制用文本提取补充
                text_extracted = ""  # 存储文本提取的版本号
                if extracted_text:  # extracted_text是页面的完整文本
                    # 文本提取正则：同样支持空格容错
                    text_version_pattern = r'(?:版\s*本|Version|rev)[:：]?\s*(V|Rev|rev)\.?\s*(\d+(?:\s*\.\s*\d+)*)'
                    text_match = re.search(text_version_pattern, extracted_text, re.IGNORECASE)
                    if text_match:
                        text_prefix = text_match.group(1).strip()
                        text_number = text_match.group(2).strip()
                        text_number = re.sub(r'\s*\.\s*', '.', text_number)  # 清理空格
                        text_extracted = f"{text_prefix}{text_number}"

                # 4. 选择最优结果：优先用“完整版本号”（如V0.1），避免短版本号
                final_version = ""
                # 规则：若表格提取不完整（长度≤2，如V0、Rev1），且文本提取更完整，则用文本提取结果
                if len(table_extracted) <= 2 and len(text_extracted) > 2:
                    final_version = text_extracted
                    # self.log_queue.put(f"版本号过短，使用文本提取结果: {final_version}\n")
                else:
                    final_version = table_extracted

                title_data["版本"] = final_version

                # TITLE提取及其他处理（保持不变）
                title_data["title"] = find_in_grid(
                    ["title","TITLE"],
                    ignore_values=ignore_list
                )
                # 无效值处理（保持不变）
                surface_treatment = title_data["表面处理"].strip().lower()
                invalid_surface_values = {"none", "无", "空白", "/", ""}
                if all(part.strip().lower() in invalid_surface_values for part in re.split(r'\s+', surface_treatment)):
                    title_data["表面处理"] = ""
                title_value = title_data["title"].strip().lower()
                invalid_title_values = {"none", "无", "空白", "/", ""}
                if all(part.strip().lower() in invalid_title_values for part in re.split(r'\s+', title_value)):
                    title_data["title"] = ""

                # 文本提取补充（同步优化加工字段）
                if extracted_text:
                    patterns = {
                        "名称": r"(?:名\s*称|Name)[:：]?\s*(\S+)",
                        "图号": r"(?:图\s*号|图\s*名|Drawing|DWG NO.)[:：]?\s*(\S+)",
                        # 加工字段文本提取模式调整
                        "加工": r"(?:加\s*工|Processing)[:：]?\s*(.*?)[\s:，;。]",
                        "材料": r"(?:材\s*料|Material)[:：]?\s*(\S+)",
                        "颜色": r"(?:颜\s*色|Color)[:：]?\s*(\S+)",
                        "表面处理": r"(?:表\s*面\s*处\s*理|Surface)[:：]?\s*(\S+)",
                        "版本": r"(?:版\s*本|Version|rev)[:：]?\s*(V|Rev|rev)\.?\s*(\d+(?:\.\d+)*)([a-zA-Z]*)",
                        "title": r"(?:title|TITLE)[:：]?\s*(\S+)"
                    }

                    # 处理加工字段的文本提取结果
                    if not title_data["加工"].strip() and "加工" in patterns:
                        match = re.search(patterns["加工"], extracted_text, re.IGNORECASE)
                        if match:
                            processing_text = match.group(1).strip()
                            # 应用相同的清洗逻辑
                            suffix_to_remove = r'(中|了|的|等|完毕|完成|结束)$'
                            while re.search(suffix_to_remove, processing_text):
                                processing_text = re.sub(suffix_to_remove, '', processing_text).strip()
                            title_data["加工"] = processing_text

                    # 其他字段处理（保持不变）
                    for key, pattern in patterns.items():
                        if key in ["表面处理", "版本", "加工"]:
                            continue
                        if key == "title" and not title_data[key].strip():
                            continue
                        if not title_data[key].strip():
                            match = re.search(pattern, extracted_text, re.IGNORECASE)
                            if match:
                                title_data[key] = match.group(1).strip()

        except Exception as e:
            print(f"提取PDF {os.path.basename(pdf_path)} 时出错: {str(e)}")

        return title_data

    def extract_excel_data_for_filling(self, excel_book, header_row, data_start_row, log_queue):
        """从Excel文件中提取数据，确保不修改表头行之前的内容"""
        # 使用xlwings处理Excel
        try:
            log_queue.put(f"表头固定在第 {header_row} 行，数据从第 {data_start_row} 行开始处理\n")

            # 获取活动工作表
            sheet = excel_book.sheets.active

            # 确定数据结束位置（最后一个有数据的行）
            last_cell = sheet.used_range.last_cell
            max_row = last_cell.row
            end_row = max_row

            # 如果没有数据行，从data_start_row开始
            if end_row < data_start_row:
                end_row = data_start_row - 1

            log_queue.put(f"检测到数据范围: 第 {data_start_row} 行至第 {end_row} 行\n")

            # 创建结果列表
            result = []

            # 读取数据
            for row in range(data_start_row, end_row + 1):
                # 物料名称 -> 配置的列
                name = str(sheet.range((row, self.name_col)).value or "").strip()

                # 物料规格 -> 配置的列
                spec = str(sheet.range((row, self.spec_col)).value or "").strip()

                # 描述 -> 配置的列
                desc = str(sheet.range((row, self.desc_col)).value or "").strip()

                # 版本 -> 配置的列
                version = str(sheet.range((row, self.version_col)).value or "").strip()

                # 新增：读取TITLE列数据
                title = str(sheet.range((row, self.title_col)).value or "").strip()

                result.append({
                    "物料名称": name,
                    "物料规格": spec,
                    "描述": desc,
                    "版本": version,
                    "title": title,  # 新增
                    "原始行号": row
                })

            return result

        except Exception as e:
            log_queue.put(f"提取Excel数据时出错: {str(e)}\n")
            return []

    def build_pdf_description(self, pdf_data):
        """构建PDF的描述字符串"""
        # 当表面处理为"无"、空白或"/"时，忽略表面处理
        include_surface = pdf_data["表面处理"] not in ["无", "空白", "/", ""]

        pdf_desc_parts = [
            pdf_data["加工"],
            pdf_data["材料"],
            pdf_data["颜色"],
            pdf_data["表面处理"] if include_surface else ""
        ]

        # 过滤空值并连接
        pdf_desc = ",".join(filter(None, [p.strip() for p in pdf_desc_parts if p.strip()]))

        # 清理多余空格和标点
        pdf_desc = re.sub(r'\s*,\s*', ',', pdf_desc).strip()
        return pdf_desc

    def fill_excel_with_pdf_data(self, excel_book, pdf_data, row_idx, name_col, spec_col, desc_col, version_col,
                                 title_col, note_start_row, log_queue):
        """将PDF数据填充到Excel的指定行，设置字体为宋体12号，新增填充TITLE列的功能"""
        try:
            # 获取活动工作表
            sheet = excel_book.sheets.active

            # 确保我们不会修改表头行之前的内容
            if row_idx <= self.header_row:
                log_queue.put(f"警告: 尝试修改表头行之前的内容，已自动调整到表头行之后\n")
                row_idx = self.header_row + 1

            log_queue.put(
                f"使用列索引: 名称={name_col}, 规格={spec_col}, 描述={desc_col}, 版本={version_col}, TITLE={title_col}\n")

            # 检查行是否存在，如果不存在，插入新行
            last_row = sheet.used_range.last_cell.row
            if row_idx > last_row:
                # 需要插入新行，按照Excel中选中表头行+1插入新行的方式
                reference_row = self.header_row + 1  # 表头行+1作为参考行

                log_queue.put(f"插入新行 {row_idx}，使用第 {reference_row} 行的格式\n")

                # 插入新行
                sheet.api.Rows(f"{row_idx}:{row_idx}").Insert()

                # 复制参考行的格式到新行
                sheet.range((reference_row, 1), (reference_row, sheet.used_range.last_cell.column)).copy()
                sheet.range((row_idx, 1)).paste(paste="formats")

            # 填充数据并设置字体为宋体12号
            updated = False

            # 设置字体格式的辅助函数
            def set_cell_value_with_font(row, col, value):
                """设置单元格值并应用字体格式，同时将半角标点转换为全角"""
                # 定义半角转全角的标点映射
                punctuation_map = {
                    ',': '，',
                    # '.': '。',
                    # ';': '；',
                    # ':': '：',
                    # '?': '？',
                    # '!': '！',
                    # '(': '（',
                    # ')': '）',
                    # '[': '［',
                    # ']': '］',
                    # '{': '｛',
                    # '}': '｝',
                    # '<': '＜',
                    # '>': '＞',
                    # '"': '“',
                    # "'": '‘',
                    # '-': '－',
                    # '_': '＿',
                    # '*': '＊',
                    # '&': '＆',
                    # '%': '％',
                    # '$': '＄',
                    # '#': '＃',
                    # '@': '＠',
                    # '^': '＾',
                    # '`': '｀'
                }

                # 如果值是字符串，进行标点转换
                if isinstance(value, str):
                    for half, full in punctuation_map.items():
                        value = value.replace(half, full)

                cell = sheet.range((row, col))
                cell.value = value
                # 设置字体为宋体，大小为12号
                cell.api.Font.Name = "宋体"
                cell.api.Font.Size = 12

            # 检查并填充名称
            if pdf_data["名称"]:
                try:
                    set_cell_value_with_font(row_idx, name_col, pdf_data["名称"])
                    updated = True
                    log_queue.put(f"已填充名称: {pdf_data['名称']} 到第 {row_idx} 行, 第 {name_col} 列\n")
                except Exception as e:
                    log_queue.put(f"填充名称时出错: {str(e)}\n")

            # 检查并填充图号
            if pdf_data["图号"]:
                try:
                    set_cell_value_with_font(row_idx, spec_col, pdf_data["图号"])
                    updated = True
                    log_queue.put(f"已填充图号: {pdf_data['图号']} 到第 {row_idx} 行, 第 {spec_col} 列\n")
                except Exception as e:
                    log_queue.put(f"填充图号时出错: {str(e)}\n")

            # 检查并填充描述
            desc = self.build_pdf_description(pdf_data)
            if desc:
                try:
                    set_cell_value_with_font(row_idx, desc_col, desc)
                    updated = True
                    log_queue.put(f"已填充描述: {desc} 到第 {row_idx} 行, 第 {desc_col} 列\n")
                except Exception as e:
                    log_queue.put(f"填充描述时出错: {str(e)}\n")

            # 检查并填充版本
            if pdf_data["版本"]:
                try:
                    set_cell_value_with_font(row_idx, version_col, pdf_data["版本"])
                    updated = True
                    log_queue.put(f"已填充版本: {pdf_data['版本']} 到第 {row_idx} 行, 第 {version_col} 列\n")
                except Exception as e:
                    log_queue.put(f"填充版本时出错: {str(e)}\n")

            # 新增：检查并填充TITLE
            if pdf_data["title"]:
                try:
                    set_cell_value_with_font(row_idx, title_col, pdf_data["title"])
                    updated = True
                    log_queue.put(f"已填充TITLE: {pdf_data['title']} 到第 {row_idx} 行, 第 {title_col} 列\n")
                except Exception as e:
                    log_queue.put(f"填充TITLE时出错: {str(e)}\n")

            return updated
        except Exception as e:
            log_queue.put(f"填充Excel时出错: {str(e)}\n")
            return False

    def process_pdf_file_for_filling(self, pdf_path, result_queue):
        """处理单个PDF文件（线程安全）"""
        try:
            pdf_data = self.extract_pdf_title_block(pdf_path)

            result = {
                "pdf_file": os.path.basename(pdf_path),
                "pdf_path": pdf_path,
                "page_count": pdf_data["页数"],
                "extracted_data": pdf_data,
                "status": "成功",
                "message": ""
            }

            # 检查是否是多页PDF
            if pdf_data["页数"] > 1:
                result["status"] = "警告"
                result["message"] = f"多页PDF ({pdf_data['页数']}页)，需要进一步查看"

            result_queue.put(result)
            return True

        except Exception as e:
            result_queue.put({
                "pdf_file": os.path.basename(pdf_path),
                "pdf_path": pdf_path,
                "page_count": 0,
                "extracted_data": {},
                "status": "错误",
                "message": f"处理错误: {str(e)}"
            })
            return False

    def process_files_for_filling(self, excel_path, pdf_folder, progress_queue, log_queue, pdf_files, excel_book,
                                  header_row, note_start_row,
                                  name_col, spec_col, desc_col, version_col, title_col):  # 新增title_col参数
        """处理PDF文件并填充到Excel中，确保只修改表头行之后的内容"""
        # 1. 提取Excel数据
        log_queue.put("读取Excel数据...\n")
        data_start_row = header_row + 1  # 数据从表头行+1开始
        excel_data = self.extract_excel_data_for_filling(excel_book, header_row, data_start_row, log_queue)

        total_pdfs = len(pdf_files)
        log_queue.put(f"开始处理 {total_pdfs} 个PDF文件...\n")

        # 2. 创建结果队列
        result_queue = queue.Queue()

        # 3. 使用线程池处理文件
        processed_count = 0
        results = []

        # 确定线程数（根据文件数量和CPU核心数）
        cpu_count = os.cpu_count() or 4
        max_workers = min(cpu_count * 2, 16, total_pdfs)

        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            # 提交所有任务
            futures = []
            for pdf_file in pdf_files:
                future = executor.submit(
                    self.process_pdf_file_for_filling,
                    pdf_file, result_queue
                )
                futures.append(future)

            # 处理结果和更新进度
            while processed_count < total_pdfs:
                try:
                    # 获取结果但不阻塞
                    if not result_queue.empty():
                        result = result_queue.get()
                        results.append(result)
                        processed_count += 1
                        progress_queue.put(processed_count)

                        # 记录处理状态
                        status_msg = f"处理: {result['pdf_file']} - {result['status']}"
                        if result['message']:
                            status_msg += f" - {result['message']}"
                        log_queue.put(status_msg + "\n")
                    else:
                        time.sleep(0.05)
                except Exception:
                    pass

        # 4. 按顺序将PDF数据填充到Excel中
        log_queue.put("将PDF数据填充到Excel中...\n")

        # 按顺序填充数据
        for i, result in enumerate(results):
            # 计算要填充的行号
            if i < len(excel_data):
                row_idx = excel_data[i]["原始行号"]
            else:
                # 如果Excel行数不足，从表头行+1开始计算新行
                row_idx = header_row + 1 + i
                log_queue.put(f"扩展Excel行到: {row_idx}\n")

            # 确保不修改表头行之前的内容
            if row_idx <= header_row:
                row_idx = header_row + 1 + i
                log_queue.put(f"自动调整行号到表头行之后: {row_idx}\n")

            # 新增：传递title_col参数
            success = self.fill_excel_with_pdf_data(excel_book, result["extracted_data"], row_idx, name_col, spec_col,
                                                    desc_col, version_col, title_col, note_start_row, log_queue)

            if not success:
                results[i]["status"] = "错误"
                results[i]["message"] = "填充Excel失败"
                log_queue.put(f"错误: 无法将 {result['pdf_file']} 的数据填充到Excel第{row_idx}行\n")
            else:
                log_queue.put(f"成功: 已将 {result['pdf_file']} 的数据填充到Excel第{row_idx}行\n")

        return results

    def generate_filling_report(self, results, output_file="./log/处理报告.txt"):
        """生成处理报告并保存到文件，包含TITLE信息"""
        with open(output_file, "w", encoding="utf-8") as f:
            f.write("=" * 70 + "\n")
            f.write("PDF信息提取并填充到Excel报告\n")
            f.write(f"生成时间: {time.strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("=" * 70 + "\n\n")

            # 统计信息
            success_count = sum(1 for r in results if r["status"] == "成功")
            warning_count = sum(1 for r in results if r["status"] == "警告")
            error_count = sum(1 for r in results if r["status"] == "错误")
            multi_page_count = sum(1 for r in results if r["page_count"] > 1)

            f.write(f"处理统计:\n")
            f.write(f"  总文件数: {len(results)}\n")
            f.write(f"  成功: {success_count}\n")
            f.write(f"  警告: {warning_count}\n")
            f.write(f"  错误: {error_count}\n")
            f.write(f"  多页PDF: {multi_page_count}\n\n")

            if multi_page_count > 0:
                f.write("需要进一步查看的多页PDF文件:\n")
                for result in results:
                    if result["page_count"] > 1:
                        f.write(f"  - {result['pdf_file']} ({result['page_count']}页)\n")
                f.write("\n")

            f.write("详细处理结果:\n\n")
            for i, result in enumerate(results, 1):
                f.write(f"文件 #{i}:\n")
                f.write(f"  文件名: {result['pdf_file']}\n")
                f.write(f"  状态: {result['status']}\n")
                if result['message']:
                    f.write(f"  消息: {result['message']}\n")
                f.write(f"  页数: {result['page_count']}\n")

                if result['extracted_data']:
                    f.write("  提取的数据:\n")
                    for key, value in result['extracted_data'].items():
                        if key != "页数":  # 页数已经单独显示
                            f.write(f"    {key}: {value}\n")

                f.write("-" * 70 + "\n")

        return os.path.abspath(output_file)

    # ======================== 比对功能函数 ========================

    def extract_excel_data(self, excel_path):
        """从Excel文件中提取申请表数据，新增提取TITLE列（第13列）"""
        try:
            # 使用openpyxl定位数据起始行
            wb = load_workbook(excel_path, data_only=True)
            ws = wb.active

            # 改进的表头查找逻辑
            start_row = 1
            header_found = False
            header_patterns = ["物料名称", "Quaero part", "物料规格", "描述", "版本", "name and specification"]

            for row_idx in range(1, 50):  # 扩大搜索范围
                row = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[row_idx]]
                if any(any(pattern in cell for pattern in header_patterns) for cell in row):
                    start_row = row_idx
                    header_found = True
                    break

            if not header_found:
                # 尝试更宽松的匹配
                for row_idx in range(1, 50):
                    row = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[row_idx]]
                    if "物料" in "".join(row) or "part" in "".join(row).lower():
                        start_row = row_idx
                        header_found = True
                        break

                if not header_found:
                    return pd.DataFrame()

            # 使用pandas读取Excel数据
            df = pd.read_excel(
                excel_path,
                header=start_row,
                keep_default_na=False
            )

            # 创建新DataFrame，只包含需要的列
            result_df = pd.DataFrame()

            # 物料名称 -> 第2列 (索引1)
            if len(df.columns) > 1:
                result_df["物料名称"] = df.iloc[:, 1].apply(lambda x: str(x).strip() if x != "" else "")
            else:
                result_df["物料名称"] = ""

            # 物料规格 -> 第3列 (索引2)
            if len(df.columns) > 2:
                result_df["物料规格"] = df.iloc[:, 2].apply(lambda x: str(x).strip() if x != "" else "")
            else:
                result_df["物料规格"] = ""

            # 描述 -> 第4列 (索引3)
            if len(df.columns) > 3:
                result_df["描述"] = df.iloc[:, 3].apply(lambda x: str(x).strip() if x != "" else "")
            else:
                result_df["描述"] = ""

            # 版本 -> 第8列 (索引7)
            if len(df.columns) > 8:
                result_df["版本"] = df.iloc[:, 8].apply(lambda x: str(x).strip() if x != "" else "")
            else:
                result_df["版本"] = ""

            # 新增：TITLE -> 第13列 (索引12)
            if len(df.columns) > 12:
                result_df["title"] = df.iloc[:, 12].apply(lambda x: str(x).strip() if x != "" else "")
            else:
                result_df["title"] = ""

            # 添加原始行号（Excel中的实际行号）
            # start_row是表头行索引，所以数据行从start_row+2开始（表头行+1）
            result_df["原始行号"] = [start_row + 2 + idx for idx in range(len(df))]

            # 移除空行
            result_df = result_df[(result_df["物料名称"] != "") | (result_df["物料规格"] != "")]

            return result_df

        except Exception as e:
            return pd.DataFrame()

    def build_excel_index(self, excel_data):
        """构建Excel数据的索引字典"""
        index = {
            "by_name": {},
            "by_spec": {},
            "by_name_spec": {},
            "by_title": {}  # 新增：按TITLE索引
        }

        for idx, row in excel_data.iterrows():
            name = row["物料名称"].strip()
            spec = row["物料规格"].strip()
            title = row["title"].strip()  # 新增
            original_row_no = row["原始行号"]

            # 添加到名称索引
            if name:
                if name not in index["by_name"]:
                    index["by_name"][name] = []
                index["by_name"][name].append((idx, original_row_no))

            # 添加到规格索引
            if spec:
                if spec not in index["by_spec"]:
                    index["by_spec"][spec] = []
                index["by_spec"][spec].append((idx, original_row_no))

            # 添加到名称+规格复合索引
            if name and spec:
                key = f"{name}|{spec}"
                if key not in index["by_name_spec"]:
                    index["by_name_spec"][key] = []
                index["by_name_spec"][key].append((idx, original_row_no))

            # 新增：添加到TITLE索引
            if title:
                if title not in index["by_title"]:
                    index["by_title"][title] = []
                index["by_title"][title].append((idx, original_row_no))

        return index

    def normalize_description(self, desc):
        """规范化描述字符串以便比较（放宽颜色对比的空格处理，保持其他部分严谨）"""
        # 1. 标准化分隔符
        desc = re.sub(r'[，；、]', ',', desc)

        # 2. 不处理材料规格的格式，保留原始空格和标点
        # 移除：desc = re.sub(r'([a-zA-Z]+\d+)\s*[,，]?\s*(t=)', r'\1 \2', desc)

        # 3. 对于颜色部分，保留原始空格，不做严格清理
        # 移除：desc = re.sub(r'\s+(?![^,]*t=)', '', desc)

        # 4. 转换为小写（仅统一大小写，其他格式保留）
        desc = desc.lower()

        # 5. 不清理多余逗号，保留原始标点
        # 移除：desc = re.sub(r',{2,}', ',', desc).strip(',')

        # 6. 保留颜色描述的处理逻辑（重新组合颜色词顺序，并去除所有空格）
        parts = desc.split(',')
        normalized_parts = []

        for part in parts:
            part = part.strip()
            if not part:
                continue

            # 检查是否是颜色描述
            if any(color in part for color in ["黑", "白", "灰", "银", "红", "蓝", "绿", "黄"]):
                part = re.sub(r'\s+\n', '\n', part)  # 只清除换行前的空格
                color_words = []
                effect_words = []

                # 常见颜色词
                color_keywords = ["黑", "白", "灰", "银", "红", "蓝", "绿", "黄"]
                # 常见效果词
                effect_keywords = ["哑光", "亮光", "磨砂", "高光", "哑", "亮", "光"]

                # 检查每个关键词
                for keyword in color_keywords + effect_keywords:
                    if keyword in part:
                        part = part.replace(keyword, "")
                        if keyword in color_keywords:
                            color_words.append(keyword)
                        else:
                            effect_words.append(keyword)

                # 重新组合颜色描述（效果词+颜色词）
                normalized_part = "".join(effect_words) + "".join(color_words)

                # 保留剩余内容（包含原有换行和合理空格）
                if part.strip():
                    normalized_part += part.strip()

                normalized_parts.append(normalized_part)
            else:
                normalized_parts.append(part)

        # 重新组合描述字符串（保留原始逗号分隔）
        desc = ','.join(normalized_parts)

        return desc

    def compare_row_with_pdf(self, excel_row, pdf_data):
        """对比Excel单行数据和PDF数据，新增对比TITLE的功能"""
        errors = []

        # 1. 物料名称对比
        excel_name = excel_row["物料名称"] if "物料名称" in excel_row else ""
        pdf_name = pdf_data["名称"]
        name_match = excel_name == pdf_name

        # 2. 物料规格对比
        excel_spec = excel_row["物料规格"] if "物料规格" in excel_row else ""
        pdf_spec = pdf_data["图号"]
        spec_match = excel_spec == pdf_spec

        # 如果名称和图号都不匹配，返回0级匹配
        if not name_match and not spec_match:
            return 0, []

        # 记录名称和图号匹配情况
        if not name_match:
            errors.append(f"物料名称不一致: Excel({excel_name}) ≠ PDF({pdf_name})")
        if not spec_match:
            errors.append(f"物料规格不一致: Excel({excel_spec}) ≠ PDF({pdf_spec})")

        # 3. 描述信息对比
        if "描述" in excel_row:
            excel_desc = excel_row["描述"]
            pdf_desc = self.build_pdf_description(pdf_data)

            # 规范化字符串
            excel_norm = self.normalize_description(excel_desc)
            pdf_norm = self.normalize_description(pdf_desc)

            # 将规范化后的描述字符串拆分成部分
            pdf_parts = [p.strip() for p in pdf_norm.split(',') if p.strip()]
            excel_parts = [p.strip() for p in excel_norm.split(',') if p.strip()]

            # 检查每个pdf_part是否在excel_parts中出现
            missing_parts = []
            for part in pdf_parts:
                if part not in excel_parts:
                    missing_parts.append(part)

            # 新增：检查每个excel_part是否在pdf_parts中出现
            extra_parts = []
            for part in excel_parts:
                if part not in pdf_parts:
                    extra_parts.append(part)

            if missing_parts:
                errors.append(f"描述不一致: Excel描述中缺少以下部分: {', '.join(missing_parts)}")

            # 新增：检查Excel描述中是否有PDF描述中没有的多余部分
            if extra_parts:
                errors.append(f"描述不一致: Excel描述中多余以下部分: {', '.join(extra_parts)}")

        # 4. 版本对比
        if "版本" in excel_row:
            excel_ver = excel_row["版本"]
            pdf_ver = pdf_data["版本"]
            if excel_ver != pdf_ver:
                errors.append(f"版本不一致: Excel({excel_ver}) ≠ PDF({pdf_ver})")

        # 新增：5. TITLE对比
        if "title" in excel_row and pdf_data.get("title"):
            excel_title = excel_row["title"]
            pdf_title = pdf_data["title"]
            if excel_title != pdf_title:
                errors.append(f"TITLE不一致: Excel({excel_title}) ≠ PDF({pdf_title})")

        # 确定匹配级别
        if not errors:
            return 2, []  # 完全匹配
        elif name_match or spec_match:
            return 1, errors  # 部分匹配
        else:
            return 0, []  # 完全不匹配

    def find_matching_rows(self, excel_data, excel_index, pdf_data):
        """使用索引查找匹配的行，新增按TITLE匹配的逻辑"""
        matches = []
        processed_rows = set()

        # 1. 尝试名称+规格完全匹配
        name = pdf_data["名称"].strip()
        spec = pdf_data["图号"].strip()
        if name and spec:
            key = f"{name}|{spec}"
            if key in excel_index["by_name_spec"]:
                for idx, row_no in excel_index["by_name_spec"][key]:
                    if idx not in processed_rows:
                        row = excel_data.loc[idx]
                        match_level, errors = self.compare_row_with_pdf(row, pdf_data)
                        matches.append((match_level, errors, idx, row_no))
                        processed_rows.add(idx)
                        # 如果是完全匹配，立即返回
                        if match_level == 2:
                            return matches

        # 2. 尝试名称匹配
        if name and name in excel_index["by_name"]:
            for idx, row_no in excel_index["by_name"][name]:
                if idx not in processed_rows:
                    row = excel_data.loc[idx]
                    match_level, errors = self.compare_row_with_pdf(row, pdf_data)
                    matches.append((match_level, errors, idx, row_no))
                    processed_rows.add(idx)
                    # 如果是完全匹配，立即返回
                    if match_level == 2:
                        return matches

        # 3. 尝试规格匹配
        if spec and spec in excel_index["by_spec"]:
            for idx, row_no in excel_index["by_spec"][spec]:
                if idx not in processed_rows:
                    row = excel_data.loc[idx]
                    match_level, errors = self.compare_row_with_pdf(row, pdf_data)
                    matches.append((match_level, errors, idx, row_no))
                    processed_rows.add(idx)
                    # 如果是完全匹配，立即返回
                    if match_level == 2:
                        return matches

        # 新增：4. 尝试TITLE匹配
        title = pdf_data["title"].strip()
        if title and title in excel_index["by_title"]:
            for idx, row_no in excel_index["by_title"][title]:
                if idx not in processed_rows:
                    row = excel_data.loc[idx]
                    match_level, errors = self.compare_row_with_pdf(row, pdf_data)
                    matches.append((match_level, errors, idx, row_no))
                    processed_rows.add(idx)
                    # 如果是完全匹配，立即返回
                    if match_level == 2:
                        return matches

        return matches

    def process_pdf_file_for_comparison(self, pdf_path, excel_data, excel_index, result_queue):
        """处理单个PDF文件（线程安全）"""
        try:
            pdf_data = self.extract_pdf_title_block(pdf_path)

            # 查找匹配的行
            matches = self.find_matching_rows(excel_data, excel_index, pdf_data)

            # 处理匹配结果
            errors = []
            found_match = False
            best_match = None
            excel_row_no = "无"  # 默认值

            # 寻找最佳匹配（最高匹配级别）
            for match in matches:
                match_level, match_errors, idx, row_no = match
                excel_row_no = row_no  # 记录行号

                if not best_match or match_level > best_match[0]:
                    best_match = (match_level, match_errors, idx, row_no)

                if match_level == 2:  # 完全匹配
                    found_match = True
                    result_queue.put((pdf_path, row_no, [], "完全匹配"))
                    break

            if not found_match and best_match:
                match_level, match_errors, idx, row_no = best_match
                if match_level == 1:  # 部分匹配
                    found_match = True
                    result_queue.put((pdf_path, row_no, match_errors, "部分匹配"))

            if not found_match:
                result_queue.put((pdf_path, "无", ["未找到匹配的Excel记录"], "无匹配"))

            return True

        except Exception as e:
            result_queue.put((pdf_path, "错误", [f"处理错误: {str(e)}"], "错误"))
            return False

    def process_files_for_comparison(self, excel_path, pdf_folder, progress_queue, log_queue, pdf_files, excel_data,
                                     excel_index):
        """优化后的文件处理函数，使用线程池管理"""
        total_pdfs = len(pdf_files)
        log_queue.put(f"开始处理 {total_pdfs} 个PDF文件...\n")

        # 创建结果队列
        result_queue = queue.Queue()

        # 使用线程池处理文件
        processed_count = 0
        results = []

        # 确定线程数（根据文件数量和CPU核心数）
        cpu_count = os.cpu_count() or 4
        max_workers = min(cpu_count * 2, 16, total_pdfs)

        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            # 提交所有任务
            futures = []
            for pdf_file in pdf_files:
                future = executor.submit(
                    self.process_pdf_file_for_comparison,
                    pdf_file, excel_data, excel_index, result_queue
                )
                futures.append(future)

            # 处理结果和更新进度
            while processed_count < total_pdfs:
                try:
                    # 获取结果但不阻塞
                    if not result_queue.empty():
                        results.append(result_queue.get())
                        processed_count += 1
                        progress_queue.put(processed_count)
                    else:
                        time.sleep(0.05)
                except Exception:
                    pass

        # 收集结果
        all_errors = []
        for pdf_path, excel_row, errors, match_type in results:
            if match_type != "完全匹配":  # 只记录有问题的匹配
                # 获取描述信息
                excel_desc = ""
                pdf_desc = ""
                excel_title = ""  # 新增
                pdf_title = ""  # 新增

                if match_type in ["部分匹配", "完全匹配"] and excel_row != "错误" and excel_row != "无":
                    try:
                        row_idx = excel_data[excel_data["原始行号"] == excel_row].index[0]
                        excel_desc = excel_data.loc[row_idx, "描述"] if "描述" in excel_data else ""
                        excel_title = excel_data.loc[row_idx, "title"] if "title" in excel_data else ""  # 新增
                    except:
                        pass

                if pdf_path != "错误":
                    try:
                        pdf_data = self.extract_pdf_title_block(pdf_path)
                        pdf_desc = self.build_pdf_description(pdf_data)
                        pdf_title = pdf_data["title"] if "title" in pdf_data else ""  # 新增
                    except:
                        pass

                all_errors.append({
                    "pdf_file": os.path.basename(pdf_path),  # 只显示文件名
                    "pdf_path": pdf_path,  # 存储完整路径
                    "excel_row": excel_row,
                    "errors": errors,
                    "excel_desc": excel_desc,
                    "pdf_desc": pdf_desc,
                    "excel_title": excel_title,  # 新增
                    "pdf_title": pdf_title,  # 新增
                    "match_type": match_type
                })

        return all_errors

    def generate_comparison_report(self, errors, output_file="./log/对比报告.txt"):
        """生成对比报告并保存到文件，包含TITLE对比信息"""
        with open(output_file, "w", encoding="utf-8") as f:
            f.write("=" * 70 + "\n")
            f.write("Excel与PDF图纸信息比对报告\n")
            f.write(f"生成时间: {time.strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("=" * 70 + "\n\n")

            if errors:
                f.write(f"发现 {len(errors)} 个错误:\n\n")
                for i, error in enumerate(errors, 1):
                    f.write(f"错误 #{i}:\n")
                    f.write(f"  PDF文件: {error['pdf_file']}\n")
                    f.write(f"  文件路径: {error['pdf_path']}\n")
                    f.write(f"  Excel行: {error['excel_row']}\n")
                    f.write(f"  匹配类型: {error['match_type']}\n")

                    # 添加描述和TITLE对比详情
                    if error['excel_desc'] or error['pdf_desc']:
                        f.write(f"  Excel描述: {error['excel_desc']}\n")
                        f.write(f"  PDF描述:  {error['pdf_desc']}\n")

                    # 新增：添加TITLE对比
                    if error['excel_title'] or error['pdf_title']:
                        f.write(f"  Excel TITLE: {error['excel_title']}\n")
                        f.write(f"  PDF TITLE:  {error['pdf_title']}\n")

                    f.write("  错误详情:\n")
                    for err_msg in error["errors"]:
                        f.write(f"    - {err_msg}\n")
                    f.write("-" * 70 + "\n")
            else:
                f.write("所有数据对比一致! 没有发现错误。\n")

        return os.path.abspath(output_file)


def main():
    """主函数"""
    root = tk.Tk()
    app = PDFExcelTool(root)
    root.mainloop()


if __name__ == "__main__":
    main()
